#!/usr/bin/env python
# coding:utf-8
'''
@author: xuyuelin
@file: test_create_log.py
@time: 2022-03-31 15:40
@desc:
'''
from iWebSNS_unittest_frameword.config.config import data_file

class Datas:
    def __init__(self):
        self.excel_path = data_file
        self.txt_path = ""
        self.yaml_path = ""

    def connect_files(self):
        import pymysql
        self.db = pymysql.connect(host='localhost', port=3306,  user='root',    password='root',
                                                        database='T107',    charset='utf-8')
        # 创建游标
        self.cur = self.db.cursor()
        return self.cur,self.db

    def read_txt(self):
        '''
        读取txt文件的内容，每一行以列表元素的形式存在
        :return: 为列表
        '''
        with open(self.txt_path,'r',encoding='utf-8') as f:
            values = f.readlines()
            f.close()
            value_list = []
            for i in range(len(values)):
                value_s = values[i].strip('\n')
                value_list.append(value_s)
        return value_list

    def write_txt(self,values):
        '''
        :param values: values参数的值为列表形式
        :return: 没有返回值
        '''
        with open(self.txt_path,'a',encoding='utf-8') as f:
            f.writelines(values)
            f.close()

    def read_excel(self,sheetname):
        '''
        :param sheetname: 传入读取的sheet的名字
        :return:  返回值是列表，每一行是列表的元素
        '''
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        table = wb[sheetname]
        rows = table.max_row
        cols = table.max_column
        list_value = []
        for row in range(2, rows + 1):
            list1 = []
            for col in range(1, cols + 1):
                col_value = table.cell(row, col).value
                list1.append(col_value)
            list_value.append(list1)
        return list_value

    def write_excel(self,*values,sheetname):
        '''
        :param values: 可变参数，可以接受多个数据，组成元组
        :param sheetname: 接受数据的表
        :return: 无返回值
        '''
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        table = wb[sheetname]
        rows = table.max_row
        values_len=len(values)
        for col in range(values_len):
            table.cell(rows+1,col+1).value = values[col]
        wb.save(self.excel_path)
        wb.close()

    def read_mysql(self,username):
        '''
        :param: username: 以用户查询数据
        :return: 数据库的数据以列表形式显示
        '''
        self.cur = self.connect_files()[0]
        sql = f'select * from users where username="{username}"'
        self.cur.execute(sql)
        values = list(self.cur.fetchall())
        return values

    def write_mysql(self,username,password):
        '''
        :param username：用户名
        :param password：密码
        :return:无返回值
        '''
        self.cur = self.connect_files()[0]
        self.db = self.connect_files()[1]
        sql = f'insert into users values("{username}","{password}")'
        self.cur.execute(sql)
        self.db.commit()

    def read_yaml(self):
        '''
        :return: 字段数据
        '''
        import yaml
        with open(self.yaml_path, 'r', encoding='utf-8') as f:
            values = f.read()
            data = yaml.safe_load(values)
        f.close()
        return data

    def write_yaml(self,data):
        '''
        :param data:字典格式的数据，或者是列表格式的数据
        :return: 无返回值
        '''
        import yaml
        with open(self.yaml_path,'a',encoding='utf-8') as f:
            yaml.safe_dump(data,f)
        f.close()


if __name__=='__main__':
    print(Datas().read_mysql('tester002'))
